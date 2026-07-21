#!/usr/bin/env python3
"""
read_package.py — Pakkemaskin Skriver (Skjåk Trelast)

Raspberry Pi lytter PASSIVT på pakkelapp-strømmen fra sorteringsanlegget
(TSX → OKI Microline) og lagrer hver pakke sikkert.

  • Knapp trykket flere ganger → DEDUP: duplikat droppes
  • Skriver av                 → tapp TX-linja; data ligger på kabelen
  • Pakke aldri kvittert        → HULL-DETEKSJON i pakkenr-rekka
  • Nullstilling (9999 → 0)     → RUNDE: oppdages automatisk, scoper dedup+hull
  • Lapp kom aldri              → MANUELL registrering (--registrer)
  • Sesong                      → rå / tørr (fysisk vribryter på maskinen)

Lappformat (dekodet fra ekte lapper):
  linje 1:  pakkenr            dimensjon (f.eks. 75X 150)
  linje 2:  dato (ÅÅÅÅ/ M/DD)  sort-siffer (5=5s, 6=krok/høgg, 4=gulv; 0 ikke i bruk)
  linje 3:                     treslag (FURU/GRAN)
  histogram (lengdefordeling, tre kolonner med antall)
  bunn:     antall · sum_lengde(1 des) · kubikk(3 des) · snittlengde(dm)
            (høyre kolonne med nuller = ubrukt, ignoreres)

Filer:
  utskrift.txt      alt råt, med tidsstempel — mister aldri noe
  pakkelapper.csv   én rad per ekte pakke (ingen duplikater)
  mangler.csv       hull i pakkenr-rekka (per runde)
  oppsummering.csv  daglig oppsummering (--oppsummering)
  pakkelapper.xlsx  Excel med ett ark per sesong (--eksporter-xlsx)
  sesong.txt        gjeldende sesong: "rå"/"tørr" (--sett-sesong)

USB-speiling (sanntid, --usb-sti):
  SD-kortet er alltid fasiten — det er der fangsten faktisk skjer, og
  ingen pakke går tapt selv om ingen minnepenn er tilkoblet. Er en
  minnepenn montert på oppgitt sti, speiles hver pakke DIT også, i
  samme øyeblikk. Er minnepennen borte når en pakke kommer inn, skrives
  den kun til SD-kortet — og neste gang minnepennen er tilkoblet
  (samme sti dukker opp igjen), synkroniseres automatisk alt som ble
  fanget i mellomtiden. Ingen manuell eksport nødvendig; bare la en
  minnepenn stå i, eller bytt den ut mot en annen når som helst.

Avhengigheter: pip install pyserial openpyxl
"""

import argparse, csv, datetime, glob, os, re, time
from pathlib import Path

FORMFEED = 0x0C
ESC = 0x1B

KOLONNER = ["tid_fanget", "dato", "pakkenr", "dimensjon", "treslag",
            "sort", "sort_navn", "antall_plank", "sum_lengde_lm", "kubikk_m3",
            "snittlengde_m", "sesong", "runde", "status", "raa"]

DATO_RE   = re.compile(r"(\d{4})\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})")
DIM_RE    = re.compile(r"(\d{2,3})\s*[xX]\s*(\d{2,4})")
TRE_RE    = re.compile(r"(FURU|GRAN)", re.I)
FLOAT1_RE = re.compile(r"(?<![\d,])(\d{1,4},\d)(?![\d])")     # nøyaktig 1 desimal
FLOAT3_RE = re.compile(r"(?<![\d,])(\d{1,4},\d{3})(?![\d])")  # nøyaktig 3 desimaler
INT_RE    = re.compile(r"(?<![\d,])(\d{1,5})(?![\d,])")       # heltall (ikke del av desimaltall)

SORT_NAVN = {
    "1": "krok",
    "3": "høgg",              # sjelden — kan forekomme, men se merknad om siffer 6 under
    "4": "gulv",              # vanlig, alltid i produksjon. Sjelden: operatør kan også
                              # bruke 4 (eller 2) for B.L — flagg de tilfellene manuelt.
    "5": "5s",
    "6": "krok",              # krok og høgg deler siffer 6 (samme spakposisjon hos sortøren).
                              # Krok er standard/flertall. Høgg (mindretall) skilles KUN ved
                              # håndskrift direkte på materialpakken — usynlig for RS-232-
                              # tappingen. Ikke et datakvalitetsproblem, bekreftet av Kent.
}

# Faneinndeling i Excel: hver bekreftet sort-kode får sin egen fane (i stedet for
# blandede dimensjoner under hverandre i én tabell). B.L (Bygningslast) og B.BL
# (Bein bygningslast) mangler ennå bekreftet siffer — se README/spør Kent.
# Pakker med ukjent/uavklart sort-siffer (2, 6, og evt. feil i 4-antagelsen) havner
# i "Uavklart" til da.
SORT_FANE = {
    "5": "5Sort",
    "1": "Krok",
    "6": "Krok",   # krok/høgg deler dette sifferet — krok er standard, se merknad i SORT_NAVN
    "3": "Hogges", # sjelden i praksis fra automatikken alene — ekte høgg skilles ved håndskrift
    "4": "Gulv",
    # "2" routes til "Uavklart" — sjeldent brukt, muligens B.L ved operatørvalg.
    # B.L kan i sjeldne tilfeller også vises som "4" — flagg de manuelt (samme
    # prinsipp som håndskrift skiller krok/høgg på siffer 6).
}
FANE_REKKEFØLGE = ["5Sort", "Krok", "Gulv", "Hogges", "B.L", "B.BL", "Uavklart"]

# Pene norske kolonnenavn for Excel (enheter i tittel, ikke gjentatt i hver celle)
KOLONNE_VISNING = {
    "tid_fanget": "Fanget", "dato": "Dato", "pakkenr": "Pakkenr",
    "dimensjon": "Dimensjon", "treslag": "Treslag", "sort": "Sort",
    "sort_navn": "Sortnavn", "antall_plank": "Antall plank",
    "sum_lengde_lm": "Sum lengde (lm)", "kubikk_m3": "Kubikk (m³)",
    "snittlengde_m": "Snittlengde (m)", "sesong": "Sesong", "runde": "Runde",
    "status": "Status", "raa": "Rådata",
    "antall_pakker": "Antall pakker", "sum_plank": "Sum plank",
    "sum_kubikk_m3": "Sum kubikk (m³)",
}


def logg(m): print(f"[{datetime.datetime.now():%H:%M:%S}] {m}", flush=True)


def rens(rabytes: bytes) -> str:
    ut, i = [], 0
    while i < len(rabytes):
        b = rabytes[i]
        if b == ESC:
            i += 2; continue
        if b in (0x0A, 0x0D):
            ut.append("\n")
        elif 0x20 <= b <= 0xFF and b != FORMFEED:
            ut.append(chr(b))
        i += 1
    linjer = [ln.rstrip() for ln in "".join(ut).split("\n")]
    return "\n".join(linjer).strip("\n")


def som_tall(s):
    try: return int(re.sub(r"\D", "", str(s)))
    except (ValueError, TypeError): return None


def _forste_ikke_null(regex, tekst):
    for m in regex.finditer(tekst):
        if float(m.group(1).replace(",", ".")) != 0:
            return m
    return None


# ── sesong ─────────────────────────────────────────────────────────
def gjett_sesong(d): return "rå" if d.month in (12, 1, 2, 3, 4, 5) else "tørr"


def les_sesong(sti: Path):
    if sti.exists():
        v = sti.read_text(encoding="utf-8").strip().lower()
        if v.startswith("rå") or v in ("raa", "ra"):    return "rå"
        if v.startswith("tør") or v in ("torr", "tor"): return "tørr"
    return gjett_sesong(datetime.date.today())


def sett_sesong(verdi, sti: Path):
    v = verdi.strip().lower()
    norm = "rå" if v in ("rå", "raa", "ra") else "tørr" if v in ("tørr", "torr", "tor") else None
    if not norm:
        logg('Bruk: --sett-sesong rå  |  --sett-sesong tørr'); return
    sti.write_text(norm, encoding="utf-8")
    logg(f"Sesong satt til: {norm}  (lagret i {sti.name})")


def parse_lapp(tekst: str, sesong: str) -> dict:
    rad = {k: "" for k in KOLONNER}
    rad["tid_fanget"] = datetime.datetime.now().isoformat(timespec="seconds")
    rad["raa"] = tekst
    rad["sesong"] = sesong

    # dato
    md = DATO_RE.search(tekst)
    if md:
        rad["dato"] = f"{md.group(1)}-{int(md.group(2)):02d}-{int(md.group(3)):02d}"
    else:
        rad["dato"] = datetime.date.today().isoformat()

    # dimensjon
    mdim = DIM_RE.search(tekst)
    if mdim:
        rad["dimensjon"] = f"{int(mdim.group(1))}x{int(mdim.group(2))}"

    # treslag
    mt = TRE_RE.search(tekst)
    if mt:
        rad["treslag"] = mt.group(1).upper()

    # sort = ensifret tall på dato-linja
    if md:
        for linje in tekst.splitlines():
            if md.group(0) in linje:
                ensifret = re.findall(r"(?<!\d)(\d)(?!\d)", DATO_RE.sub(" ", linje))
                if ensifret:
                    rad["sort"] = ensifret[0]
                break
    rad["sort_navn"] = SORT_NAVN.get(rad["sort"], "")

    # pakkenr = første heltall når dato og dimensjon er fjernet
    uten = tekst
    if md:   uten = uten.replace(md.group(0), " ")
    if mdim: uten = uten.replace(mdim.group(0), " ")
    mpk = INT_RE.search(uten)
    if mpk:
        rad["pakkenr"] = mpk.group(1)

    # summeringsblokk, ankret på desimaltall
    m1 = _forste_ikke_null(FLOAT1_RE, tekst)   # sum_lengde (1 desimal)
    m3 = _forste_ikke_null(FLOAT3_RE, tekst)   # kubikk (3 desimaler)
    if m1: rad["sum_lengde_lm"] = m1.group(1).replace(",", ".")
    if m3: rad["kubikk_m3"]    = m3.group(1).replace(",", ".")

    # antall = siste IKKE-NULL heltall før sum_lengde
    if m1:
        før = [t for t in INT_RE.findall(tekst[:m1.start()]) if int(t) != 0]
        if før:
            rad["antall_plank"] = før[-1]
    # snittlengde = første ikke-null heltall etter kubikk (dm → m)
    if m3:
        for t in INT_RE.findall(tekst[m3.end():]):
            if int(t) != 0:
                rad["snittlengde_m"] = f"{int(t)/10:.1f}"
                break
    return rad


class Register:
    """Dedup + hull-deteksjon, scopet per RUNDE. Ny runde oppdages når
    pakkenr hopper langt bakover (9999 → 0). Rekonstrueres fra CSV."""
    def __init__(self, csv_sti: Path, mangler_sti: Path, terskel: int = 100):
        self.csv_sti, self.mangler_sti, self.terskel = csv_sti, mangler_sti, terskel
        self.runde, self.sett, self.maks = 1, set(), None
        if csv_sti.exists():
            runder = {}
            with csv_sti.open(encoding="utf-8") as f:
                for rad in csv.DictReader(f):
                    if rad.get("status") not in ("ok", "manuell"):
                        continue
                    r = som_tall(rad.get("runde")) or 1
                    n = som_tall(rad.get("pakkenr"))
                    if n is not None:
                        runder.setdefault(r, set()).add(n)
            if runder:
                self.runde = max(runder)
                self.sett = runder[self.runde]
                self.maks = max(self.sett)
            logg(f"Lastet runde {self.runde}: {len(self.sett)} pakkenr (høyeste {self.maks}).")

    def vurder(self, pakkenr_tekst):
        n = som_tall(pakkenr_tekst)
        if n is None:
            return "ukjent", None, False
        if self.maks is not None and (self.maks - n) > self.terskel:
            return "ok", n, True                      # stort hopp bakover = nullstilling
        if n in self.sett:
            return "duplikat", n, False
        return "ok", n, False

    def ny_runde(self):
        self.runde += 1
        self.sett, self.maks = set(), None
        logg(f"↻ Ny runde {self.runde} — pakkenr ser nullstilt ut (9999 → 0)")

    def registrer(self, n):
        if self.maks is not None and n > self.maks + 1:
            self._skriv_hull(self.maks + 1, n - 1)
        self.sett.add(n)
        self.maks = n if self.maks is None else max(self.maks, n)

    def _skriv_hull(self, fra, til):
        ny = not self.mangler_sti.exists()
        with self.mangler_sti.open("a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if ny: w.writerow(["oppdaget", "runde", "manglende_pakkenr", "merknad"])
            for m in range(fra, til + 1):
                w.writerow([datetime.datetime.now().isoformat(timespec="seconds"), self.runde, m,
                            "mulig hull — sjekk om pakken ble kvittert"])
        logg(f"⚠  Mulig hull (runde {self.runde}): pakkenr {fra}–{til} → {self.mangler_sti.name}")


def append_csv(rad, csv_sti: Path):
    ny = not csv_sti.exists()
    with csv_sti.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=KOLONNER, extrasaction="ignore")
        if ny: w.writeheader()
        w.writerow(rad)


# ── USB-speiling (sanntid, med automatisk innhenting) ────────────────
def _usb_tilgjengelig(usb_sti: Path | None) -> bool:
    """Sjekker om minnepennen er montert akkurat nå (kan komme/gå når som helst)."""
    return usb_sti is not None and usb_sti.parent.exists()


def _nokkel(rad) -> tuple:
    """(runde, pakkenr) brukes til å finne ut hvilke rader som allerede
    finnes på minnepennen, slik at vi ikke dupliserer ved synkronisering."""
    return (str(rad.get("runde") or ""), str(rad.get("pakkenr") or ""))


def _les_nokler(csv_sti: Path) -> set:
    if not csv_sti.exists():
        return set()
    with csv_sti.open(encoding="utf-8") as f:
        return {_nokkel(rad) for rad in csv.DictReader(f)}


def synkroniser_usb(csv_sti: Path, usb_sti: Path | None):
    """Kalles hver gang minnepennen er tilkoblet. Sammenligner hva som
    finnes på SD-kortet mot hva som finnes på minnepennen, og etterfyller
    automatisk alt som ble fanget mens minnepennen var borte (f.eks. mens
    den forrige turen til PC-en pågikk). Trygt å kalle ofte — den gjør
    ingenting hvis minnepennen allerede er fullt oppdatert."""
    if not _usb_tilgjengelig(usb_sti) or not csv_sti.exists():
        return
    try:
        nokler_usb = _les_nokler(usb_sti)
        manglet = 0
        with csv_sti.open(encoding="utf-8") as f:
            for rad in csv.DictReader(f):
                if _nokkel(rad) not in nokler_usb:
                    append_csv(rad, usb_sti)
                    manglet += 1
        if manglet:
            logg(f"📀 Minnepenn oppdatert — hentet inn {manglet} pakke(r) som ble fanget mens den var frakoblet.")
    except OSError as e:
        logg(f"⚠  Klarte ikke synkronisere minnepenn ({e}).")


def skriv_utskrift(tekst, utskrift: Path):
    with utskrift.open("a", encoding="utf-8") as f:
        f.write(f"\n===== {datetime.datetime.now().isoformat(timespec='seconds')} =====\n{tekst}\n")


def behandle(tekst, utskrift, csv_sti, reg: Register, sesong, bare_fangst, usb_sti: Path | None = None):
    if not tekst: return
    skriv_utskrift(tekst, utskrift)
    if bare_fangst:
        logg("rå lapp fanget"); return
    pakkenr = parse_lapp(tekst, sesong)["pakkenr"]
    status, n, reset = reg.vurder(pakkenr)
    if status == "duplikat":
        logg(f"↺ duplikat pakkenr {pakkenr} — droppet (finnes i utskrift.txt)"); return
    if reset:
        reg.ny_runde()
    rad = parse_lapp(tekst, sesong)
    rad["runde"] = reg.runde
    rad["status"] = status
    if status == "ok":
        reg.registrer(n)
    append_csv(rad, csv_sti)                 # SD-kortet — alltid, er fasiten
    synkroniser_usb(csv_sti, usb_sti)         # speiler denne pakken + evt. "hull" fra forrige frakobling, i ett steg
    if status == "ukjent":
        logg("? fant ikke pakkenr — lagret med rådata for manuell sjekk")
    else:
        usb_status = " 📀" if _usb_tilgjengelig(usb_sti) else ""
        logg(f"✓ pakke {rad['pakkenr']} [r{reg.runde}/{sesong}]{usb_status} "
             f"{rad['dimensjon'] or '?'} {rad['treslag'] or '?'} "
             f"sort {rad['sort'] or '?'}({rad['sort_navn'] or '?'}), "
             f"{rad['antall_plank'] or '?'} plank, {rad['kubikk_m3'] or '?'} m³")


def registrer_manuelt(pakkenr, csv_sti, mangler_sti, sesong, terskel):
    reg = Register(csv_sti, mangler_sti, terskel)
    status, n, reset = reg.vurder(pakkenr)
    if status == "duplikat":
        logg(f"Pakkenr {pakkenr} finnes allerede i runde {reg.runde}. Avbryter."); return
    if n is None:
        logg("Ugyldig pakkenr."); return
    if reset:
        reg.ny_runde()
    reg.registrer(n)
    rad = {k: "" for k in KOLONNER}
    rad.update(tid_fanget=datetime.datetime.now().isoformat(timespec="seconds"),
               dato=datetime.date.today().isoformat(), pakkenr=str(n),
               sesong=sesong, runde=reg.runde, status="manuell", raa="(manuelt registrert)")
    append_csv(rad, csv_sti)
    logg(f"✓ Manuelt registrert pakke {n} [r{reg.runde}/{sesong}]")


def _tid_aggreger(csv_sti):
    """Leser CSV én gang og bøtter ok/manuell-pakker på år, måned (ÅÅÅÅ-MM),
    ISO-uke (ÅÅÅÅ-UNN) og dag. Tar med pakker, plank, kubikk og løpemeter.
    Skiller også på sesong (rå/tørr) per bøtte, for produksjonssammenligning.
    Returnerer (per_ar, per_mnd, per_uke, per_dag) — hver er dict nøkkel->tall."""
    def ny():
        return {"pakker": 0, "plank": 0, "kubikk": 0.0, "lm": 0.0, "rå": 0, "tørr": 0}
    per_ar, per_mnd, per_uke, per_dag = {}, {}, {}, {}
    if not csv_sti.exists():
        return per_ar, per_mnd, per_uke, per_dag
    with csv_sti.open(encoding="utf-8") as f:
        for rad in csv.DictReader(f):
            if rad.get("status") not in ("ok", "manuell"):
                continue
            try:
                d = datetime.date.fromisoformat((rad.get("dato") or "").strip())
            except ValueError:
                continue
            plank = som_tall(rad.get("antall_plank")) or 0
            try: kubikk = float(str(rad.get("kubikk_m3", "")).replace(",", "."))
            except ValueError: kubikk = 0.0
            try: lm = float(str(rad.get("sum_lengde_lm", "")).replace(",", "."))
            except ValueError: lm = 0.0
            sesong = (rad.get("sesong") or "").strip()
            iso = d.isocalendar()
            for bøtte, nøkkel in ((per_ar, str(d.year)),
                                  (per_mnd, f"{d.year}-{d.month:02d}"),
                                  (per_uke, f"{iso[0]}-U{iso[1]:02d}"),
                                  (per_dag, d.isoformat())):
                g = bøtte.setdefault(nøkkel, ny())
                g["pakker"] += 1
                g["plank"] += plank
                g["kubikk"] += kubikk
                g["lm"] += lm
                if sesong in ("rå", "tørr"):
                    g[sesong] += 1
    return per_ar, per_mnd, per_uke, per_dag


def _grupper_rader(csv_sti, nokkel_felter):
    """Leser pakkelapper.csv og grupperer ok/manuell-rader på de gitte feltene.
    Returnerer dict: tuple(nøkkelverdier) -> {pakker, plank, kubikk}."""
    grupper = {}
    if not csv_sti.exists():
        return grupper
    with csv_sti.open(encoding="utf-8") as f:
        for rad in csv.DictReader(f):
            if rad.get("status") not in ("ok", "manuell"):
                continue
            nøkkel = tuple(rad.get(felt) or "ukjent" for felt in nokkel_felter)
            g = grupper.setdefault(nøkkel, {"pakker": 0, "plank": 0, "kubikk": 0.0})
            g["pakker"] += 1
            g["plank"] += som_tall(rad.get("antall_plank")) or 0
            try: g["kubikk"] += float(str(rad.get("kubikk_m3", "")).replace(",", "."))
            except ValueError: pass
    return grupper


def oppsummering(csv_sti, ut_sti):
    if not csv_sti.exists():
        logg(f"Finner ikke {csv_sti}."); return
    grupper = _grupper_rader(csv_sti, ["sesong", "dato"])
    with ut_sti.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sesong", "dato", "antall_pakker", "sum_plank", "sum_kubikk_m3"])
        for (sesong, dato) in sorted(grupper):
            g = grupper[(sesong, dato)]
            w.writerow([sesong, dato, g["pakker"], g["plank"], round(g["kubikk"], 3)])
    logg(f"Oppsummering skrevet til {ut_sti.name}")
    for (sesong, dato) in sorted(grupper):
        g = grupper[(sesong, dato)]
        print(f"   {sesong:5} {dato}:  {g['pakker']} pakker,  {g['plank']} plank,  {round(g['kubikk'],3)} m³")


def oppsummering_dimensjon(csv_sti, ut_sti):
    """Summerer pakker/plank/kubikk gruppert på sesong + dimensjon (uavhengig av dato)."""
    if not csv_sti.exists():
        logg(f"Finner ikke {csv_sti}."); return
    grupper = _grupper_rader(csv_sti, ["sesong", "dimensjon"])
    if not grupper:
        logg("Ingen pakker å summere ennå."); return
    with ut_sti.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sesong", "dimensjon", "antall_pakker", "sum_plank", "sum_kubikk_m3"])
        for (sesong, dim) in sorted(grupper):
            g = grupper[(sesong, dim)]
            w.writerow([sesong, dim, g["pakker"], g["plank"], round(g["kubikk"], 3)])
    logg(f"Oversikt per dimensjon skrevet til {ut_sti.name}")
    for (sesong, dim) in sorted(grupper):
        g = grupper[(sesong, dim)]
        print(f"   {sesong:5} {dim:>9}:  {g['pakker']} pakker,  {g['plank']} plank,  {round(g['kubikk'],3)} m³")


def list_porter():
    try:
        from serial.tools import list_ports
        porter = [p.device for p in list_ports.comports()]
    except Exception:
        porter = sorted(glob.glob("/dev/ttyUSB*") + glob.glob("/dev/ttyACM*"))
    print("Serieporter funnet:" if porter else "Fant ingen serieporter.")
    for p in porter: print("  ", p)


def les_serie(args, utskrift, csv_sti, reg, sesong, usb_sti: Path | None = None):
    import serial
    paritet = {"N": serial.PARITY_NONE, "E": serial.PARITY_EVEN, "O": serial.PARITY_ODD}[args.paritet]
    logg(f"Starter. Lytter på {args.port} @ {args.baud} {args.databits}{args.paritet}{args.stoppbits}"
         f"  sesong={sesong}{'  [BARE FANGST]' if args.bare_fangst else ''}"
         f"{f'  usb={usb_sti}' if usb_sti else ''}  (Ctrl+C for å stoppe)")
    while True:
        try:
            ser = serial.Serial(args.port, args.baud, bytesize=args.databits,
                                parity=paritet, stopbits=args.stoppbits, timeout=args.timeout)
        except serial.SerialException as e:
            logg(f"Får ikke åpnet {args.port} ({e}). Prøver igjen om 5 s …"); time.sleep(5); continue
        logg("Tilkoblet.")
        buf, sist = bytearray(), time.time()
        try:
            while True:
                b = ser.read(1)
                if b:
                    if b[0] == FORMFEED:
                        behandle(rens(bytes(buf)), utskrift, csv_sti, reg, sesong, args.bare_fangst, usb_sti); buf.clear()
                    else:
                        buf += b
                    sist = time.time()
                elif buf and (time.time() - sist) > args.flush:
                    behandle(rens(bytes(buf)), utskrift, csv_sti, reg, sesong, args.bare_fangst, usb_sti); buf.clear()
        except serial.SerialException as e:
            logg(f"Mistet forbindelsen ({e}). Kobler til igjen om 5 s …")
            try: ser.close()
            except Exception: pass
            time.sleep(5); continue
        except KeyboardInterrupt:
            if buf: behandle(rens(bytes(buf)), utskrift, csv_sti, reg, sesong, args.bare_fangst, usb_sti)
            ser.close(); logg("Stoppet."); return


def kjor_simulering(args, utskrift, csv_sti, reg, sesong, usb_sti: Path | None = None):
    data = Path(args.simuler).read_bytes()
    logg(f"Simulerer fra {args.simuler} …  sesong={sesong}")
    for chunk in data.split(bytes([FORMFEED])):
        behandle(rens(chunk), utskrift, csv_sti, reg, sesong, args.bare_fangst, usb_sti)
    logg("Ferdig.")


def eksporter_xlsx(csv_sti, xlsx_sti):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not csv_sti.exists():
        logg(f"Finner ikke {csv_sti}."); return

    FONT_NAVN = "Calibri"
    TEMA_GRØNN = "1B4D3E"      # tittel/header-bakgrunn — varm, dempet, trelast-aktig
    GRÅ_TEKST = "595959"
    BAND_GRÅ = "F2F2F2"        # svak skygge — markerer ny dimensjon, ikke en sterk farge
    KANT = Side(style="thin", color="D9D9D9")
    RAMME = Border(left=KANT, right=KANT, top=KANT, bottom=KANT)
    TITTEL_FONT = Font(name=FONT_NAVN, size=14, bold=True, color="FFFFFF")
    UNDERTITTEL_FONT = Font(name=FONT_NAVN, size=9, italic=True, color="D9D9D9")
    HEADER_FONT = Font(name=FONT_NAVN, size=10, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor=TEMA_GRØNN)
    TITTEL_FILL = PatternFill("solid", fgColor=TEMA_GRØNN)
    DATA_FONT = Font(name=FONT_NAVN, size=10)
    RAA_FONT = Font(name=FONT_NAVN, size=8, italic=True, color=GRÅ_TEKST)
    SUBTOTAL_FONT = Font(name=FONT_NAVN, size=9, bold=True, color="1B4D3E")
    SUBTOTAL_FILL = PatternFill("solid", fgColor="E8F0EE")

    # Sort: kun brukt i "Uavklart"-fanen (der flere sorter blandes); i de navngitte
    # fanene (5Sort/Krok/Hogges) er sorten allerede gitt av fanenavnet, så fargekode
    # på sort-cellen er overflødig der.
    SORT_FARGER = {
        "5": (PatternFill("solid", fgColor="C6EFCE"), Font(name=FONT_NAVN, size=10, color="2E7D32")),
        "4": (PatternFill("solid", fgColor="BDD7EE"), Font(name=FONT_NAVN, size=10, color="1F4E78")),
        "1": (PatternFill("solid", fgColor="FFEB9C"), Font(name=FONT_NAVN, size=10, color="9C6500")),
        "3": (PatternFill("solid", fgColor="F8CBAD"), Font(name=FONT_NAVN, size=10, color="943126")),
        "6": (PatternFill("solid", fgColor="E0E0E0"), Font(name=FONT_NAVN, size=10, color="595959")),
    }

    NUMMER_FORMAT = {
        "antall_plank": "#,##0", "sum_lengde_lm": "#,##0.0", "kubikk_m3": "0.000",
        "snittlengde_m": "0.0", "runde": "0", "pakkenr": "0",
        "antall_pakker": "#,##0", "sum_plank": "#,##0", "sum_kubikk_m3": "0.000",
    }
    # Bredde = plass til hele overskriften + buffer til nedtrekkspilen fra autofilter,
    # ellers kuttes teksten visuelt bak pilen (f.eks. "Antall pl..").
    KOLONNEBREDDE = {
        "tid_fanget": 18, "dato": 13, "pakkenr": 11, "dimensjon": 13, "treslag": 12,
        "sort": 8, "sort_navn": 14, "antall_plank": 16, "sum_lengde_lm": 18,
        "kubikk_m3": 15, "snittlengde_m": 18, "sesong": 11, "runde": 9,
        "status": 11, "raa": 40,
        "antall_pakker": 16, "sum_plank": 13, "sum_kubikk_m3": 18,
    }

    def skriv_header(ws, kol_start, rad_nr, felter):
        for j, felt in enumerate(felter, start=kol_start):
            c = ws.cell(rad_nr, j, KOLONNE_VISNING.get(felt, felt))
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[rad_nr].height = 30

    def skriv_dataceller(ws, rad_nr, kol_start, felter, rad, skyggelagt):
        for j, felt in enumerate(felter, start=kol_start):
            verdi = rad.get(felt, "")
            if felt == "dato" and verdi:
                try: verdi = datetime.date.fromisoformat(verdi)
                except ValueError: pass
            elif felt == "tid_fanget" and verdi:
                try: verdi = datetime.datetime.fromisoformat(verdi)
                except ValueError: pass
            elif felt in ("antall_plank", "kubikk_m3", "sum_lengde_lm", "snittlengde_m",
                          "runde", "pakkenr", "antall_pakker", "sum_plank", "sum_kubikk_m3"):
                try: verdi = float(verdi) if verdi != "" else None
                except (ValueError, TypeError): pass
            c = ws.cell(rad_nr, j, verdi)
            c.border = RAMME
            c.font = RAA_FONT if felt == "raa" else DATA_FONT
            if felt == "dato": c.number_format = "DD.MM.YYYY"
            elif felt == "tid_fanget": c.number_format = "DD.MM.YYYY HH:MM"
            elif felt in NUMMER_FORMAT: c.number_format = NUMMER_FORMAT[felt]
            if felt in ("dimensjon", "sesong", "status"):
                c.alignment = Alignment(horizontal="center")
            if felt in ("sort", "sort_navn"):
                c.alignment = Alignment(horizontal="center")
                fyll = SORT_FARGER.get(rad.get("sort"))
                if fyll: c.fill, c.font = fyll
            if skyggelagt and felt != "raa" and not (felt in ("sort", "sort_navn") and rad.get("sort") in SORT_FARGER):
                c.fill = PatternFill("solid", fgColor=BAND_GRÅ)

    def skriv_fane(ws, tittel, felter, rader, vis_dimensjonsoversikt=True):
        n = len(felter)
        bredde_kol = max(n, 5)
        ws.sheet_view.showGridLines = False
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=bredde_kol)
        c = ws.cell(1, 1, tittel); c.font = TITTEL_FONT; c.fill = TITTEL_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=bredde_kol)
        c = ws.cell(2, 1, f"Generert {datetime.datetime.now():%d.%m.%Y %H:%M}  ·  {len(rader)} rader")
        c.font = UNDERTITTEL_FONT; c.fill = TITTEL_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for j in range(1, bredde_kol + 1):
            felt_her = felter[j - 1] if j <= n else ""
            ws.column_dimensions[get_column_letter(j)].width = KOLONNEBREDDE.get(felt_her, 12)

        rad_nr = 4

        # Liten dimensjonsoversikt øverst i fanen — egen, kort tabell (ikke filtrerbar,
        # så den forstyrrer ikke autofilteret på hovedtabellen under).
        if vis_dimensjonsoversikt and rader:
            dim_sum = {}
            for r in rader:
                d = r.get("dimensjon") or "ukjent"
                g = dim_sum.setdefault(d, {"pakker": 0, "plank": 0, "kubikk": 0.0})
                g["pakker"] += 1
                g["plank"] += som_tall(r.get("antall_plank")) or 0
                try: g["kubikk"] += float(str(r.get("kubikk_m3", "")).replace(",", "."))
                except ValueError: pass
            c = ws.cell(rad_nr, 1, "Per dimensjon i denne fanen:")
            c.font = SUBTOTAL_FONT
            rad_nr += 1
            overskrift = ["Dimensjon", "Pakker", "Plank", "Kubikk (m³)"]
            for j, h in enumerate(overskrift, start=1):
                c = ws.cell(rad_nr, j, h); c.font = SUBTOTAL_FONT; c.fill = SUBTOTAL_FILL
                c.alignment = Alignment(horizontal="center")
            rad_nr += 1
            for d in sorted(dim_sum):
                g = dim_sum[d]
                verdier = [d, g["pakker"], g["plank"], round(g["kubikk"], 3)]
                for j, v in enumerate(verdier, start=1):
                    c = ws.cell(rad_nr, j, v); c.fill = SUBTOTAL_FILL
                    c.font = Font(name=FONT_NAVN, size=9)
                    c.alignment = Alignment(horizontal="center")
                    if j == 4: c.number_format = "0.000"
                rad_nr += 1
            rad_nr += 1  # luft før hovedtabellen

        header_rad = rad_nr
        skriv_header(ws, 1, header_rad, felter)
        forrige_dim = None
        skygge = False
        for i, rad in enumerate(rader, start=header_rad + 1):
            if rad.get("dimensjon") != forrige_dim:
                skygge = not skygge
                forrige_dim = rad.get("dimensjon")
            skriv_dataceller(ws, i, 1, felter, rad, skygge)

        ws.freeze_panes = f"A{header_rad + 1}"
        siste_kol = get_column_letter(n)
        ws.auto_filter.ref = f"A{header_rad}:{siste_kol}{header_rad + len(rader)}"

    def skriv_sammendrag(ws):
        """Egen 'Sammendrag'-fane øverst i arbeidsboka: totaler per år, måned,
        uke og dag — så ekspeditøren slipper å regne sammen selv."""
        per_ar, per_mnd, per_uke, per_dag = _tid_aggreger(csv_sti)
        ws.sheet_view.showGridLines = False
        BREDDE = 6
        for j, b in enumerate([16, 13, 13, 15, 15, 20], start=1):
            ws.column_dimensions[get_column_letter(j)].width = b

        # Tittelbånd
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=BREDDE)
        c0 = ws.cell(1, 1, "Pakkelapper — sammendrag"); c0.font = TITTEL_FONT; c0.fill = TITTEL_FILL
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=BREDDE)
        c0 = ws.cell(2, 1, f"Generert {datetime.datetime.now():%d.%m.%Y %H:%M}  ·  alt regnet ut automatisk")
        c0.font = UNDERTITTEL_FONT; c0.fill = TITTEL_FILL
        c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        rad = [4]  # muterbar teller så indre funksjon kan øke den

        def blokk(tittel, data_dict, nyeste_forst=True, maks=None):
            r = rad[0]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=BREDDE)
            ct = ws.cell(r, 1, tittel); ct.font = Font(name=FONT_NAVN, size=11, bold=True, color="FFFFFF")
            ct.fill = HEADER_FILL; ct.alignment = Alignment(horizontal="left", indent=1)
            r += 1
            overskrifter = ["Periode", "Pakker", "Plank", "Kubikk (m³)", "Løpemeter", "Sesong (rå/tørr)"]
            for j, h in enumerate(overskrifter, start=1):
                c = ws.cell(r, j, h); c.font = SUBTOTAL_FONT; c.fill = SUBTOTAL_FILL
                c.alignment = Alignment(horizontal="center")
            r += 1
            nøkler = sorted(data_dict.keys(), reverse=nyeste_forst)
            if maks:
                nøkler = nøkler[:maks]
            for nøkkel in nøkler:
                g = data_dict[nøkkel]
                verdier = [nøkkel, g["pakker"], g["plank"], round(g["kubikk"], 3),
                           round(g["lm"], 1), f'{g["rå"]} / {g["tørr"]}']
                for j, v in enumerate(verdier, start=1):
                    c = ws.cell(r, j, v)
                    c.font = DATA_FONT
                    c.border = RAMME
                    if j == 1:   c.alignment = Alignment(horizontal="center")
                    if j == 4:   c.number_format = "0.000"
                    if j == 5:   c.number_format = "#,##0.0"
                    if j == 6:   c.alignment = Alignment(horizontal="center")
                r += 1
            rad[0] = r + 1  # luft før neste blokk

        if not per_ar:
            ws.cell(4, 1, "Ingen pakker registrert ennå.").font = DATA_FONT
            return
        blokk("PER ÅR", per_ar)
        blokk("PER MÅNED", per_mnd)
        blokk("PER UKE (siste 12)", per_uke, maks=12)
        blokk("PER DAG (siste 21)", per_dag, maks=21)

        # Nøkkeltall nederst
        r = rad[0]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=BREDDE)
        ct = ws.cell(r, 1, "NØKKELTALL"); ct.font = Font(name=FONT_NAVN, size=11, bold=True, color="FFFFFF")
        ct.fill = HEADER_FILL; ct.alignment = Alignment(horizontal="left", indent=1)
        r += 1
        i_ar = str(datetime.date.today().year)
        dager_i_ar = [v for k, v in per_dag.items() if k.startswith(i_ar)]
        snitt = round(sum(d["pakker"] for d in dager_i_ar) / len(dager_i_ar), 1) if dager_i_ar else 0
        beste = max(dager_i_ar, key=lambda d: d["pakker"], default=None)
        beste_dag = max((k for k in per_dag if k.startswith(i_ar)),
                        key=lambda k: per_dag[k]["pakker"], default="—")
        nøkkeltall = [
            (f"Snitt pakker per produksjonsdag ({i_ar})", snitt),
            ("Beste enkeltdag i år", f'{beste_dag}: {beste["pakker"]} pakker' if beste else "—"),
            ("Antall produksjonsdager i år", len(dager_i_ar)),
        ]
        for tittel, verdi in nøkkeltall:
            ws.cell(r, 1, tittel).font = DATA_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws.cell(r, 4, verdi); c.font = SUBTOTAL_FONT
            r += 1
        ws.freeze_panes = "A3"

    wb = Workbook()
    with csv_sti.open(encoding="utf-8") as f:
        alle_rader = list(csv.DictReader(f))

    fane_data = {navn: [] for navn in FANE_REKKEFØLGE}
    for rad in alle_rader:
        fane = SORT_FANE.get(rad.get("sort"), "Uavklart")
        fane_data[fane].append(rad)
    for rader in fane_data.values():
        rader.sort(key=lambda r: (r.get("dimensjon") or "", r.get("dato") or "", som_tall(r.get("pakkenr")) or 0))

    FELTER_KATEGORI = [f for f in KOLONNER if f not in ("sort", "sort_navn")]   # sort er gitt av fanen
    FELTER_UAVKLART = KOLONNER                                                  # her trengs sort/sortnavn

    # Sammendrag alltid først — det ekspeditøren ser når fila åpnes
    sammendrag_ws = wb.active
    sammendrag_ws.title = "Sammendrag"
    skriv_sammendrag(sammendrag_ws)

    forste = True
    for fane in FANE_REKKEFØLGE:
        rader = fane_data[fane]
        if not rader: continue
        ws = wb.create_sheet()
        ws.title = fane
        forste = False
        felter = FELTER_UAVKLART if fane == "Uavklart" else FELTER_KATEGORI
        skriv_fane(ws, f"Pakkelapper — {fane}", felter, rader)
    if forste:
        tom_ws = wb.create_sheet()
        tom_ws.title = "5Sort"
        skriv_fane(tom_ws, "Pakkelapper — 5Sort", FELTER_KATEGORI, [])

    # Samlet oversikt: antall/plank/kubikk per dimensjon, på tvers av alle fanene
    dim_grupper = _grupper_rader(csv_sti, ["sesong", "dimensjon"])
    if dim_grupper:
        dim_felter = ["sesong", "dimensjon", "antall_pakker", "sum_plank", "sum_kubikk_m3"]
        dim_rader = [
            {"sesong": s, "dimensjon": d, "antall_pakker": g["pakker"],
             "sum_plank": g["plank"], "sum_kubikk_m3": round(g["kubikk"], 3)}
            for (s, d), g in sorted(dim_grupper.items())
        ]
        dim_ws = wb.create_sheet("Per dimensjon (alle)")
        skriv_fane(dim_ws, "Oversikt per dimensjon — alle sorter", dim_felter, dim_rader,
                   vis_dimensjonsoversikt=False)

    tmp = xlsx_sti.with_suffix(".tmp.xlsx")
    wb.save(tmp); os.replace(tmp, xlsx_sti)
    antall = {fane: len(rader) for fane, rader in fane_data.items() if rader}
    logg(f"Eksportert til {xlsx_sti}  (" + ", ".join(f"{f}: {n}" for f, n in antall.items()) + ")")



def main():
    p = argparse.ArgumentParser(description="Pakkemaskin Skriver — RS-232 → CSV/Excel.")
    p.add_argument("--port", default="/dev/ttyUSB0")
    p.add_argument("--baud", type=int, default=9600)
    p.add_argument("--databits", type=int, default=8)
    p.add_argument("--paritet", choices=["N", "E", "O"], default="N")
    p.add_argument("--stoppbits", type=int, default=1)
    p.add_argument("--timeout", type=float, default=1.0)
    p.add_argument("--flush", type=float, default=3.0)
    p.add_argument("--reset-terskel", type=int, default=100,
                   help="hvor stort hopp bakover som regnes som nullstilling")
    p.add_argument("--csv", default="pakkelapper.csv")
    p.add_argument("--xlsx", default="pakkelapper.xlsx")
    p.add_argument("--utskrift", default="utskrift.txt")
    p.add_argument("--mangler", default="mangler.csv")
    p.add_argument("--oppsummering-fil", default="oppsummering.csv")
    p.add_argument("--dimensjon-fil", default="oppsummering_dimensjon.csv")
    p.add_argument("--sesong-fil", default="sesong.txt")
    p.add_argument("--usb-sti", default="/media/usb0",
                   help='mappe der en minnepenn forventes montert, for sanntids-speiling av '
                        'pakkelapper.csv. SD-kortet er alltid fasiten uansett — sett til "" '
                        'for å slå speiling helt av.')
    p.add_argument("--bare-fangst", action="store_true")
    p.add_argument("--simuler")
    p.add_argument("--list-porter", action="store_true")
    p.add_argument("--eksporter-xlsx", action="store_true")
    p.add_argument("--oppsummering", action="store_true")
    p.add_argument("--oppsummering-dimensjon", action="store_true",
                   help="antall/plank/kubikk gruppert per dimensjon (uavhengig av dato)")
    p.add_argument("--registrer", help="registrer et pakkenr manuelt (når lappen aldri kom)")
    p.add_argument("--sett-sesong", help='sett gjeldende sesong: "rå" eller "tørr"')
    args = p.parse_args()

    utskrift, csv_sti = Path(args.utskrift), Path(args.csv)
    mangler_sti, xlsx_sti, sesong_fil = Path(args.mangler), Path(args.xlsx), Path(args.sesong_fil)
    usb_sti = (Path(args.usb_sti) / args.csv) if args.usb_sti else None

    if args.list_porter:    list_porter(); return
    if args.sett_sesong:    sett_sesong(args.sett_sesong, sesong_fil); return
    if args.eksporter_xlsx: eksporter_xlsx(csv_sti, xlsx_sti); return
    if args.oppsummering:   oppsummering(csv_sti, Path(args.oppsummering_fil)); return
    if args.oppsummering_dimensjon: oppsummering_dimensjon(csv_sti, Path(args.dimensjon_fil)); return

    sesong = les_sesong(sesong_fil)
    if args.registrer:
        registrer_manuelt(args.registrer, csv_sti, mangler_sti, sesong, args.reset_terskel)
        synkroniser_usb(csv_sti, usb_sti)
        return

    reg = Register(csv_sti, mangler_sti, args.reset_terskel)
    if _usb_tilgjengelig(usb_sti):
        logg(f"Minnepenn funnet ved oppstart ({usb_sti.parent}) — sjekker om noe mangler …")
        synkroniser_usb(csv_sti, usb_sti)
    if args.simuler: kjor_simulering(args, utskrift, csv_sti, reg, sesong, usb_sti)
    else:            les_serie(args, utskrift, csv_sti, reg, sesong, usb_sti)


if __name__ == "__main__":
    main()
