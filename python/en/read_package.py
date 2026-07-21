#!/usr/bin/env python3
"""
les_pakkelapp.py — rs232excel - Package label capture (Timber sawmill)

Raspberry Pi lytter PASSIVT på pakkelapp-strømmen fra sorteringsanlegget
(TSX → OKI Microline) og lagrer hver pakke sikkert.

  • Knapp trykket flere ganger → DEDUP: duplikat droppes
  • Skriver av                 → tapp TX-linja; data ligger på kabelen
  • Pakke aldri kvittert        → HULL-DETEKSJON i pakkenr-rekka
  • Nullstilling (9999 → 0)     → RUNDE: oppdages automatisk, scoper dedup+hull
  • Lapp kom aldri              → MANUELL registrering (--register)
  • Sesong                      → raw / kiln-dried (fysisk vribryter på maskinen)

Lappformat (dekodet fra ekte lapper):
  linje 1:  pakkenr            dimensjon (f.eks. 75X 150)
  linje 2:  dato (ÅÅÅÅ/ M/DD)  sort-siffer (5=5s, 6=krok/høgg, 4=gulv; 0 ikke i bruk)
  linje 3:                     treslag (FURU/GRAN)
  histogram (lengdefordeling, tre kolonner med antall)
  bunn:     antall · sum_lengde(1 des) · kubikk(3 des) · snittlengde(dm)
            (høyre kolonne med nuller = ubrukt, ignoreres)

Filer:
  capture.txt      alt rawt, med tidsstempel — mister aldri noe
  packages.csv   én rad per ekte pakke (ingen duplikater)
  missing.csv       hull i pakkenr-rekka (per runde)
  oppsummering.csv  daglig oppsummering (--summary)
  pakkelapper.xlsx  Excel med ett ark per sesong (--export-xlsx)
  season.txt        gjeldende sesong: "raw"/"kiln-dried" (--set-season)

Avhengigheter: pip install pyserial openpyxl
"""

import argparse, csv, datetime, glob, os, re, time
from pathlib import Path

FORMFEED = 0x0C
ESC = 0x1B

KOLONNER = ["tid_fanget", "dato", "pakkenr", "dimensjon", "treslag",
            "sort", "sort_navn", "antall_plank", "sum_lengde_lm", "kubikk_m3",
            "snittlengde_m", "sesong", "runde", "status", "raa"]

DATO_RE   = re.compile(r"(\d{4})\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})")
DIM_RE    = re.compile(r"(\d{2,3})\s*[xX]\s*(\d{2,4})")
TRE_RE    = re.compile(r"(FURU|GRAN)", re.I)
FLOAT1_RE = re.compile(r"(?<![\d,])(\d{1,4},\d)(?![\d])")     # nøyaktig 1 desimal
FLOAT3_RE = re.compile(r"(?<![\d,])(\d{1,4},\d{3})(?![\d])")  # nøyaktig 3 desimaler
INT_RE    = re.compile(r"(?<![\d,])(\d{1,5})(?![\d,])")       # heltall (ikke del av desimaltall)

SORT_NAVN = {
    "1": "crooked",
    "3": "rejected",            # rare from PLC alone — true rejected often marked by hand
    "4": "floor",
    "5": "5th grade",
    "6": "crooked",             # crooked and rejected share digit 6 at the sorter
}

# Excel tabs: one sheet per confirmed grade code.
SORT_FANE = {
    "5": "5th Grade",
    "1": "Crooked",
    "6": "Crooked",
    "3": "Rejected",
    "4": "Floor",
}
FANE_REKKEFØLGE = ["5th Grade", "Crooked", "Floor", "Rejected", "B.L", "B.BL", "No Category"]
FANE_UAVKLART = "No Category"

# Column headers for Excel (units in title, not repeated in every cell)
KOLONNE_VISNING = {
    "tid_fanget": "Captured", "dato": "Date", "pakkenr": "Package no",
    "dimensjon": "Dimension", "treslag": "Species", "sort": "Grade code",
    "sort_navn": "Grade", "antall_plank": "Board count",
    "sum_lengde_lm": "Total length (lm)", "kubikk_m3": "Volume (m³)",
    "snittlengde_m": "Avg length (m)", "sesong": "Season", "runde": "Round",
    "status": "Status", "raa": "Raw line",
    "antall_pakker": "Packages", "sum_plank": "Boards",
    "sum_kubikk_m3": "Volume (m³)",
}


def log(m): print(f"[{datetime.datetime.now():%H:%M:%S}] {m}", flush=True)


# ESC-sekvenser (Epson FX / IBM-emulering, som OKI ML280 Elite bruker):
# verdi = antall parameterbytes ETTER kommandotegnet.
#   "NUL" = les til NUL-byte (tabstopp-lister)
#   "GFX" = n1 n2 + (n1 + 256*n2) databytes (punktgrafikk)
# Ukjente kommandoer: antar 0 parametre og loger, så ekte --raw-capture
# avslører dem i stedet for at parametre lekker inn i lappteksten.
ESC_PARAM = {
    "@":0,"E":0,"F":0,"G":0,"H":0,"M":0,"P":0,"T":0,"<":0,
    "0":0,"1":0,"2":0,"4":0,"5":0,"6":0,"7":0,"8":0,"9":0,
    "x":1,"U":1,"W":1,"S":1,"J":1,"3":1,"A":1,"l":1,"Q":1,"N":1,
    "R":1,"k":1,"m":1,"t":1,"I":1,"i":1,"a":1,"-":1,"p":1,"r":1,
    "s":1,"j":1,"!":1,"+":1,"C":1,   # ESC C NUL n håndteres spesielt under
    "D":"NUL","B":"NUL",
    "K":"GFX","L":"GFX","Y":"GFX","Z":"GFX",
}

def clean_text(rabytes: bytes) -> str:
    ut, i = [], 0
    n = len(rabytes)
    while i < n:
        b = rabytes[i]
        if b == ESC:
            if i + 1 >= n:
                break
            cmd = chr(rabytes[i + 1])
            spec = ESC_PARAM.get(cmd)
            i += 2                                  # ESC + kommandotegn
            if spec is None:
                log(f"?  ukjent ESC-sekvens: ESC {cmd!r} — antar ingen parametre")
            elif spec == "NUL":
                while i < n and rabytes[i] != 0x00:
                    i += 1
                i += 1                              # selve NUL-byten
            elif spec == "GFX":
                if i + 2 <= n:
                    antall = rabytes[i] + 256 * rabytes[i + 1]
                    i += 2 + antall
            elif cmd == "C" and i < n and rabytes[i] == 0x00:
                i += 2                              # ESC C NUL n (sidelengde i tommer)
            else:
                i += spec
            continue
        if b in (0x0A, 0x0D):
            ut.append("\n")
        elif 0x20 <= b <= 0xFF and b != FORMFEED:
            ut.append(chr(b))
        i += 1
    linjer = [ln.rstrip() for ln in "".join(ut).split("\n")]
    return "\n".join(linjer).strip("\n")


def as_int(s):
    try: return int(re.sub(r"\D", "", str(s)))
    except (ValueError, TypeError): return None


def _forste_ikke_null(regex, tekst):
    for m in regex.finditer(tekst):
        if float(m.group(1).replace(",", ".")) != 0:
            return m
    return None


# ── sesong ─────────────────────────────────────────────────────────
def gjett_sesong(d): return "raw" if d.month in (12, 1, 2, 3, 4, 5) else "kiln-dried"


def read_season(sti: Path):
    if sti.exists():
        v = sti.read_text(encoding="utf-8").strip().lower()
        if v.startswith("raw") or v in ("raa", "ra"):    return "raw"
        if v.startswith("tør") or v in ("torr", "tor"): return "kiln-dried"
    return gjett_sesong(datetime.date.today())


def set_season(verdi, sti: Path):
    v = verdi.strip().lower()
    norm = "raw" if v in ("raw", "raa", "ra") else "kiln-dried" if v in ("kiln-dried", "torr", "tor") else None
    if not norm:
        log('Bruk: --set-season raw  |  --set-season kiln-dried'); return
    sti.write_text(norm, encoding="utf-8")
    log(f"Sesong satt til: {norm}  (lagret i {sti.name})")


def parse_label(tekst: str, sesong: str) -> dict:
    rad = {k: "" for k in KOLONNER}
    rad["tid_fanget"] = datetime.datetime.now().isoformat(timespec="seconds")
    rad["raa"] = tekst
    rad["sesong"] = sesong

    # dato
    md = DATO_RE.search(tekst)
    if md:
        rad["dato"] = f"{md.group(1)}-{int(md.group(2)):02d}-{int(md.group(3)):02d}"
    else:
        rad["dato"] = datetime.date.today().isoformat()

    # dimensjon
    mdim = DIM_RE.search(tekst)
    if mdim:
        rad["dimensjon"] = f"{int(mdim.group(1))}x{int(mdim.group(2))}"

    # treslag
    mt = TRE_RE.search(tekst)
    if mt:
        rad["treslag"] = mt.group(1).upper()

    # sort = ensifret tall på dato-linja
    if md:
        for linje in tekst.splitlines():
            if md.group(0) in linje:
                ensifret = re.findall(r"(?<!\d)(\d)(?!\d)", DATO_RE.sub(" ", linje))
                if ensifret:
                    rad["sort"] = ensifret[0]
                break
    rad["sort_navn"] = SORT_NAVN.get(rad["sort"], "")

    # pakkenr = første heltall når dato og dimensjon er fjernet
    uten = tekst
    if md:   uten = uten.replace(md.group(0), " ")
    if mdim: uten = uten.replace(mdim.group(0), " ")
    mpk = INT_RE.search(uten)
    if mpk:
        rad["pakkenr"] = mpk.group(1)

    # summeringsblokk, ankret på desimaltall
    m1 = _forste_ikke_null(FLOAT1_RE, tekst)   # sum_lengde (1 desimal)
    m3 = _forste_ikke_null(FLOAT3_RE, tekst)   # kubikk (3 desimaler)
    if m1: rad["sum_lengde_lm"] = m1.group(1).replace(",", ".")
    if m3: rad["kubikk_m3"]    = m3.group(1).replace(",", ".")

    # antall = siste IKKE-NULL heltall før sum_lengde
    if m1:
        før = [t for t in INT_RE.findall(tekst[:m1.start()]) if int(t) != 0]
        if før:
            rad["antall_plank"] = før[-1]
    # snittlengde = første ikke-null heltall etter kubikk (dm → m)
    if m3:
        for t in INT_RE.findall(tekst[m3.end():]):
            if int(t) != 0:
                rad["snittlengde_m"] = f"{int(t)/10:.1f}"
                break
    return rad


class Register:
    """Dedup + hull-deteksjon, scopet per RUNDE. Ny runde oppdages når
    pakkenr hopper langt bakover (9999 → 0). Rekonstrueres fra CSV."""
    def __init__(self, csv_sti: Path, mangler_sti: Path, terskel: int = 100):
        self.csv_sti, self.mangler_sti, self.terskel = csv_sti, mangler_sti, terskel
        self.runde, self.sett, self.maks = 1, set(), None
        if csv_sti.exists():
            runder = {}
            with csv_sti.open(encoding="utf-8") as f:
                for rad in csv.DictReader(f):
                    if rad.get("status") not in ("ok", "manuell"):
                        continue
                    r = as_int(rad.get("runde")) or 1
                    n = as_int(rad.get("pakkenr"))
                    if n is not None:
                        runder.setdefault(r, set()).add(n)
            if runder:
                self.runde = max(runder)
                self.sett = runder[self.runde]
                self.maks = max(self.sett)
            log(f"Lastet runde {self.runde}: {len(self.sett)} pakkenr (høyeste {self.maks}).")

    def vurder(self, pakkenr_tekst):
        n = as_int(pakkenr_tekst)
        if n is None:
            return "ukjent", None, False
        if self.maks is not None and (self.maks - n) > self.terskel:
            return "ok", n, True                      # stort hopp bakover = nullstilling
        if n in self.sett:
            return "duplikat", n, False
        return "ok", n, False

    def ny_runde(self):
        self.runde += 1
        self.sett, self.maks = set(), None
        log(f"↻ Ny runde {self.runde} — pakkenr ser nullstilt ut (9999 → 0)")

    def registrer(self, n):
        if self.maks is not None and n > self.maks + 1:
            self._skriv_hull(self.maks + 1, n - 1)
        self.sett.add(n)
        self.maks = n if self.maks is None else max(self.maks, n)

    def _skriv_hull(self, fra, til):
        ny = not self.mangler_sti.exists()
        with self.mangler_sti.open("a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if ny: w.writerow(["oppdaget", "runde", "manglende_pakkenr", "merknad"])
            for m in range(fra, til + 1):
                w.writerow([datetime.datetime.now().isoformat(timespec="seconds"), self.runde, m,
                            "mulig hull — sjekk om pakken ble kvittert"])
        log(f"⚠  Mulig hull (runde {self.runde}): pakkenr {fra}–{til} → {self.mangler_sti.name}")


def append_csv(rad, csv_sti: Path):
    ny = not csv_sti.exists()
    with csv_sti.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=KOLONNER, extrasaction="ignore")
        if ny: w.writeheader()
        w.writerow(rad)


def skriv_utskrift(tekst, utskrift: Path):
    with utskrift.open("a", encoding="utf-8") as f:
        f.write(f"\n===== {datetime.datetime.now().isoformat(timespec='seconds')} =====\n{tekst}\n")


def behandle(tekst, utskrift, csv_sti, reg: Register, sesong, bare_fangst):
    if not tekst: return
    skriv_utskrift(tekst, utskrift)
    if bare_fangst:
        log("raw lapp fanget"); return
    pakkenr = parse_label(tekst, sesong)["pakkenr"]
    status, n, reset = reg.vurder(pakkenr)
    if status == "duplikat":
        log(f"↺ duplikat pakkenr {pakkenr} — droppet (finnes i capture.txt)"); return
    if reset:
        reg.ny_runde()
    rad = parse_label(tekst, sesong)
    rad["runde"] = reg.runde
    rad["status"] = status
    if status == "ok":
        reg.registrer(n)
    append_csv(rad, csv_sti)
    if status == "ukjent":
        log("? fant ikke pakkenr — lagret med rawdata for manuell sjekk")
    else:
        log(f"✓ pakke {rad['pakkenr']} [r{reg.runde}/{sesong}] "
             f"{rad['dimensjon'] or '?'} {rad['treslag'] or '?'} "
             f"sort {rad['sort'] or '?'}({rad['sort_navn'] or '?'}), "
             f"{rad['antall_plank'] or '?'} plank, {rad['kubikk_m3'] or '?'} m³")


def registrer_manuelt(pakkenr, csv_sti, mangler_sti, sesong, terskel):
    reg = Register(csv_sti, mangler_sti, terskel)
    status, n, reset = reg.vurder(pakkenr)
    if status == "duplikat":
        log(f"Pakkenr {pakkenr} finnes allerede i runde {reg.runde}. Avbryter."); return
    if n is None:
        log("Ugyldig pakkenr."); return
    if reset:
        reg.ny_runde()
    reg.registrer(n)
    rad = {k: "" for k in KOLONNER}
    rad.update(tid_fanget=datetime.datetime.now().isoformat(timespec="seconds"),
               dato=datetime.date.today().isoformat(), pakkenr=str(n),
               sesong=sesong, runde=reg.runde, status="manuell", raa="(manuelt registrert)")
    append_csv(rad, csv_sti)
    log(f"✓ Manuelt registrert pakke {n} [r{reg.runde}/{sesong}]")


def _grupper_rader(csv_sti, nokkel_felter):
    """Leser packages.csv og grupperer ok/manuell-rader på de gitte feltene.
    Returnerer dict: tuple(nøkkelverdier) -> {pakker, plank, kubikk}."""
    grupper = {}
    if not csv_sti.exists():
        return grupper
    with csv_sti.open(encoding="utf-8") as f:
        for rad in csv.DictReader(f):
            if rad.get("status") not in ("ok", "manuell"):
                continue
            nøkkel = tuple(rad.get(felt) or "ukjent" for felt in nokkel_felter)
            g = grupper.setdefault(nøkkel, {"pakker": 0, "plank": 0, "kubikk": 0.0})
            g["pakker"] += 1
            g["plank"] += as_int(rad.get("antall_plank")) or 0
            try: g["kubikk"] += float(str(rad.get("kubikk_m3", "")).replace(",", "."))
            except ValueError: pass
    return grupper


def oppsummering(csv_sti, ut_sti):
    if not csv_sti.exists():
        log(f"Finner ikke {csv_sti}."); return
    grupper = _grupper_rader(csv_sti, ["sesong", "dato"])
    with ut_sti.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sesong", "dato", "antall_pakker", "sum_plank", "sum_kubikk_m3"])
        for (sesong, dato) in sorted(grupper):
            g = grupper[(sesong, dato)]
            w.writerow([sesong, dato, g["pakker"], g["plank"], round(g["kubikk"], 3)])
    log(f"Oppsummering skrevet til {ut_sti.name}")
    for (sesong, dato) in sorted(grupper):
        g = grupper[(sesong, dato)]
        print(f"   {sesong:5} {dato}:  {g['pakker']} pakker,  {g['plank']} plank,  {round(g['kubikk'],3)} m³")


def oppsummering_dimensjon(csv_sti, ut_sti):
    """Summerer pakker/plank/kubikk gruppert på sesong + dimensjon (uavhengig av dato)."""
    if not csv_sti.exists():
        log(f"Finner ikke {csv_sti}."); return
    grupper = _grupper_rader(csv_sti, ["sesong", "dimensjon"])
    if not grupper:
        log("Ingen pakker å summere ennå."); return
    with ut_sti.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sesong", "dimensjon", "antall_pakker", "sum_plank", "sum_kubikk_m3"])
        for (sesong, dim) in sorted(grupper):
            g = grupper[(sesong, dim)]
            w.writerow([sesong, dim, g["pakker"], g["plank"], round(g["kubikk"], 3)])
    log(f"Oversikt per dimensjon skrevet til {ut_sti.name}")
    for (sesong, dim) in sorted(grupper):
        g = grupper[(sesong, dim)]
        print(f"   {sesong:5} {dim:>9}:  {g['pakker']} pakker,  {g['plank']} plank,  {round(g['kubikk'],3)} m³")


def list_porter():
    try:
        from serial.tools import list_ports
        porter = [p.device for p in list_ports.comports()]
    except Exception:
        porter = sorted(glob.glob("/dev/ttyUSB*") + glob.glob("/dev/ttyACM*"))
    print("Serieporter funnet:" if porter else "Fant ingen serieporter.")
    for p in porter: print("  ", p)


def les_serie(args, utskrift, csv_sti, reg, sesong):
    import serial
    paritet = {"N": serial.PARITY_NONE, "E": serial.PARITY_EVEN, "O": serial.PARITY_ODD}[args.parity]
    log(f"Listening on {args.port} @ {args.baud} {args.databits}{args.parity}{args.stoppbits}"
         f"  season={sesong}{'  [RAW CAPTURE]' if args.raw_capture else ''}  (Ctrl+C to stop)")
    while True:
        try:
            ser = serial.Serial(args.port, args.baud, bytesize=args.databits,
                                parity=paritet, stopbits=args.stoppbits, timeout=args.timeout)
        except serial.SerialException as e:
            log(f"Får ikke åpnet {args.port} ({e}). Prøver igjen om 5 s …"); time.sleep(5); continue
        log("Tilkoblet.")
        buf, sist = bytearray(), time.time()
        try:
            while True:
                b = ser.read(1)
                if b:
                    if b[0] == FORMFEED:
                        behandle(clean_text(bytes(buf)), utskrift, csv_sti, reg, sesong, args.raw_capture); buf.clear()
                    else:
                        buf += b
                    sist = time.time()
                elif buf and (time.time() - sist) > args.flush:
                    behandle(clean_text(bytes(buf)), utskrift, csv_sti, reg, sesong, args.raw_capture); buf.clear()
        except serial.SerialException as e:
            log(f"Lost connection ({e}). Reconnecting in 5 s …")
            try: ser.close()
            except Exception: pass
            time.sleep(5); continue
        except KeyboardInterrupt:
            if buf: behandle(clean_text(bytes(buf)), utskrift, csv_sti, reg, sesong, args.raw_capture)
            ser.close(); log("Stopped."); return


def kjor_simulering(args, utskrift, csv_sti, reg, sesong):
    data = Path(args.simulate).read_bytes()
    log(f"Simulating from {args.simulate} …  season={sesong}")
    for chunk in data.split(bytes([FORMFEED])):
        behandle(clean_text(chunk), utskrift, csv_sti, reg, sesong, args.raw_capture)
    log("Done.")


def eksporter_xlsx(csv_sti, xlsx_sti):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not csv_sti.exists():
        log(f"Finner ikke {csv_sti}."); return

    FONT_NAVN = "Calibri"
    TEMA_GRØNN = "1B4D3E"      # tittel/header-bakgrunn — varm, dempet, trelast-aktig
    GRÅ_TEKST = "595959"
    BAND_GRÅ = "F2F2F2"        # svak skygge — markerer ny dimensjon, ikke en sterk farge
    KANT = Side(style="thin", color="D9D9D9")
    RAMME = Border(left=KANT, right=KANT, top=KANT, bottom=KANT)
    TITTEL_FONT = Font(name=FONT_NAVN, size=14, bold=True, color="FFFFFF")
    UNDERTITTEL_FONT = Font(name=FONT_NAVN, size=9, italic=True, color="D9D9D9")
    HEADER_FONT = Font(name=FONT_NAVN, size=10, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor=TEMA_GRØNN)
    TITTEL_FILL = PatternFill("solid", fgColor=TEMA_GRØNN)
    DATA_FONT = Font(name=FONT_NAVN, size=10)
    RAA_FONT = Font(name=FONT_NAVN, size=8, italic=True, color=GRÅ_TEKST)
    SUBTOTAL_FONT = Font(name=FONT_NAVN, size=9, bold=True, color="1B4D3E")
    SUBTOTAL_FILL = PatternFill("solid", fgColor="E8F0EE")

    # Sort: kun brukt i "Uavklart"-fanen (der flere sorter blandes); i de navngitte
    # fanene (5Sort/Krok/Hogges) er sorten allerede gitt av fanenavnet, så fargekode
    # på sort-cellen er overflødig der.
    SORT_FARGER = {
        "5": (PatternFill("solid", fgColor="C6EFCE"), Font(name=FONT_NAVN, size=10, color="2E7D32")),
        "4": (PatternFill("solid", fgColor="BDD7EE"), Font(name=FONT_NAVN, size=10, color="1F4E78")),
        "1": (PatternFill("solid", fgColor="FFEB9C"), Font(name=FONT_NAVN, size=10, color="9C6500")),
        "3": (PatternFill("solid", fgColor="F8CBAD"), Font(name=FONT_NAVN, size=10, color="943126")),
        "6": (PatternFill("solid", fgColor="E0E0E0"), Font(name=FONT_NAVN, size=10, color="595959")),
    }

    NUMMER_FORMAT = {
        "antall_plank": "#,##0", "sum_lengde_lm": "#,##0.0", "kubikk_m3": "0.000",
        "snittlengde_m": "0.0", "runde": "0", "pakkenr": "0",
        "antall_pakker": "#,##0", "sum_plank": "#,##0", "sum_kubikk_m3": "0.000",
    }
    # Bredde = plass til hele overskriften + buffer til nedtrekkspilen fra autofilter,
    # ellers kuttes teksten visuelt bak pilen (f.eks. "Antall pl..").
    KOLONNEBREDDE = {
        "tid_fanget": 18, "dato": 13, "pakkenr": 11, "dimensjon": 13, "treslag": 12,
        "sort": 8, "sort_navn": 14, "antall_plank": 16, "sum_lengde_lm": 18,
        "kubikk_m3": 15, "snittlengde_m": 18, "sesong": 11, "runde": 9,
        "status": 11, "raa": 40,
        "antall_pakker": 16, "sum_plank": 13, "sum_kubikk_m3": 18,
    }

    def skriv_header(ws, kol_start, rad_nr, felter):
        for j, felt in enumerate(felter, start=kol_start):
            c = ws.cell(rad_nr, j, KOLONNE_VISNING.get(felt, felt))
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[rad_nr].height = 30

    def skriv_dataceller(ws, rad_nr, kol_start, felter, rad, skyggelagt):
        for j, felt in enumerate(felter, start=kol_start):
            verdi = rad.get(felt, "")
            if felt == "dato" and verdi:
                try: verdi = datetime.date.fromisoformat(verdi)
                except ValueError: pass
            elif felt == "tid_fanget" and verdi:
                try: verdi = datetime.datetime.fromisoformat(verdi)
                except ValueError: pass
            elif felt in ("antall_plank", "kubikk_m3", "sum_lengde_lm", "snittlengde_m",
                          "runde", "pakkenr", "antall_pakker", "sum_plank", "sum_kubikk_m3"):
                try: verdi = float(verdi) if verdi != "" else None
                except (ValueError, TypeError): pass
            c = ws.cell(rad_nr, j, verdi)
            c.border = RAMME
            c.font = RAA_FONT if felt == "raa" else DATA_FONT
            if felt == "dato": c.number_format = "DD.MM.YYYY"
            elif felt == "tid_fanget": c.number_format = "DD.MM.YYYY HH:MM"
            elif felt in NUMMER_FORMAT: c.number_format = NUMMER_FORMAT[felt]
            if felt in ("dimensjon", "sesong", "status"):
                c.alignment = Alignment(horizontal="center")
            if felt in ("sort", "sort_navn"):
                c.alignment = Alignment(horizontal="center")
                fyll = SORT_FARGER.get(rad.get("sort"))
                if fyll: c.fill, c.font = fyll
            if skyggelagt and felt != "raa" and not (felt in ("sort", "sort_navn") and rad.get("sort") in SORT_FARGER):
                c.fill = PatternFill("solid", fgColor=BAND_GRÅ)

    def skriv_fane(ws, tittel, felter, rader, vis_dimensjonsoversikt=True):
        n = len(felter)
        bredde_kol = max(n, 5)
        ws.sheet_view.showGridLines = False
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=bredde_kol)
        c = ws.cell(1, 1, tittel); c.font = TITTEL_FONT; c.fill = TITTEL_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=bredde_kol)
        c = ws.cell(2, 1, f"Generated {datetime.datetime.now():%d.%m.%Y %H:%M}  ·  {len(rader)} rows")
        c.font = UNDERTITTEL_FONT; c.fill = TITTEL_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for j in range(1, bredde_kol + 1):
            felt_her = felter[j - 1] if j <= n else ""
            ws.column_dimensions[get_column_letter(j)].width = KOLONNEBREDDE.get(felt_her, 12)

        rad_nr = 4

        # Liten dimensjonsoversikt øverst i fanen — egen, kort tabell (ikke filtrerbar,
        # så den forstyrrer ikke autofilteret på hovedtabellen under).
        if vis_dimensjonsoversikt and rader:
            dim_sum = {}
            for r in rader:
                d = r.get("dimensjon") or "unknown"
                g = dim_sum.setdefault(d, {"pakker": 0, "plank": 0, "kubikk": 0.0})
                g["pakker"] += 1
                g["plank"] += as_int(r.get("antall_plank")) or 0
                try: g["kubikk"] += float(str(r.get("kubikk_m3", "")).replace(",", "."))
                except ValueError: pass
            c = ws.cell(rad_nr, 1, "Per dimension in this sheet:")
            c.font = SUBTOTAL_FONT
            rad_nr += 1
            overskrift = ["Dimension", "Packages", "Boards", "Volume (m³)"]
            for j, h in enumerate(overskrift, start=1):
                c = ws.cell(rad_nr, j, h); c.font = SUBTOTAL_FONT; c.fill = SUBTOTAL_FILL
                c.alignment = Alignment(horizontal="center")
            rad_nr += 1
            for d in sorted(dim_sum):
                g = dim_sum[d]
                verdier = [d, g["pakker"], g["plank"], round(g["kubikk"], 3)]
                for j, v in enumerate(verdier, start=1):
                    c = ws.cell(rad_nr, j, v); c.fill = SUBTOTAL_FILL
                    c.font = Font(name=FONT_NAVN, size=9)
                    c.alignment = Alignment(horizontal="center")
                    if j == 4: c.number_format = "0.000"
                rad_nr += 1
            rad_nr += 1  # luft før hovedtabellen

        header_rad = rad_nr
        skriv_header(ws, 1, header_rad, felter)
        forrige_dim = None
        skygge = False
        for i, rad in enumerate(rader, start=header_rad + 1):
            if rad.get("dimensjon") != forrige_dim:
                skygge = not skygge
                forrige_dim = rad.get("dimensjon")
            skriv_dataceller(ws, i, 1, felter, rad, skygge)

        ws.freeze_panes = f"A{header_rad + 1}"
        siste_kol = get_column_letter(n)
        ws.auto_filter.ref = f"A{header_rad}:{siste_kol}{header_rad + len(rader)}"

    wb = Workbook()
    with csv_sti.open(encoding="utf-8") as f:
        alle_rader = list(csv.DictReader(f))

    fane_data = {navn: [] for navn in FANE_REKKEFØLGE}
    for rad in alle_rader:
        fane = SORT_FANE.get(rad.get("sort"), FANE_UAVKLART)
        fane_data[fane].append(rad)
    for rader in fane_data.values():
        rader.sort(key=lambda r: (r.get("dimensjon") or "", r.get("dato") or "", as_int(r.get("pakkenr")) or 0))

    FELTER_KATEGORI = [f for f in KOLONNER if f not in ("sort", "sort_navn")]   # sort er gitt av fanen
    FELTER_UAVKLART = KOLONNER                                                  # her trengs sort/sortnavn

    forste = True
    for fane in FANE_REKKEFØLGE:
        rader = fane_data[fane]
        if not rader: continue
        ws = wb.active if forste else wb.create_sheet()
        ws.title = fane
        forste = False
        felter = FELTER_UAVKLART if fane == FANE_UAVKLART else FELTER_KATEGORI
        skriv_fane(ws, f"Package labels — {fane}", felter, rader)
    if forste:
        wb.active.title = "5th Grade"
        skriv_fane(wb.active, "Package labels — 5th Grade", FELTER_KATEGORI, [])

    # Samlet oversikt: antall/plank/kubikk per dimensjon, på tvers av alle fanene
    dim_grupper = _grupper_rader(csv_sti, ["sesong", "dimensjon"])
    if dim_grupper:
        dim_felter = ["sesong", "dimensjon", "antall_pakker", "sum_plank", "sum_kubikk_m3"]
        dim_rader = [
            {"sesong": s, "dimensjon": d, "antall_pakker": g["pakker"],
             "sum_plank": g["plank"], "sum_kubikk_m3": round(g["kubikk"], 3)}
            for (s, d), g in sorted(dim_grupper.items())
        ]
        dim_ws = wb.create_sheet("By dimension (all)")
        skriv_fane(dim_ws, "Overview by dimension — all grades", dim_felter, dim_rader,
                   vis_dimensjonsoversikt=False)

    # ── Rådata (flatt ark) — grunnlaget Sammendrag-formlene regner på ────────
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    ok_rader = []
    for r in alle_rader:
        if r.get("status") not in ("ok", "manuell"):
            continue
        try:    d = datetime.date.fromisoformat(r.get("dato", ""))
        except ValueError: continue
        ok_rader.append({
            "dato": d,
            "kategori": SORT_FANE.get(r.get("sort"), FANE_UAVKLART),
            "sesong": r.get("sesong", ""),
            "dimensjon": r.get("dimensjon", ""),
            "plank": as_int(r.get("antall_plank")) or 0,
            "lm": float(str(r.get("sum_lengde_lm") or 0).replace(",", ".") or 0),
            "kubikk": float(str(r.get("kubikk_m3") or 0).replace(",", ".") or 0),
        })
    ok_rader.sort(key=lambda r: r["dato"])

    rd = wb.create_sheet("Raw data")
    rd.sheet_view.showGridLines = False
    for j, (felt, tittel, br) in enumerate([
            ("dato", "Date", 12), ("kategori", "Category", 12),
            ("sesong", "Season", 10), ("dimensjon", "Dimension", 12),
            ("plank", "Boards", 10), ("lm", "Running metres (lm)", 15),
            ("kubikk", "Volume (m³)", 13)], start=1):
        c = rd.cell(1, j, tittel)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        rd.column_dimensions[get_column_letter(j)].width = br
    for i, r in enumerate(ok_rader, start=2):
        rd.cell(i, 1, r["dato"]).number_format = "DD.MM.YYYY"
        rd.cell(i, 2, r["kategori"])
        rd.cell(i, 3, r["sesong"])
        rd.cell(i, 4, r["dimensjon"])
        rd.cell(i, 5, r["plank"]).number_format = "#,##0"
        rd.cell(i, 6, round(r["lm"], 1)).number_format = "#,##0.0"
        rd.cell(i, 7, round(r["kubikk"], 3)).number_format = "0.000"
        for j in range(1, 8):
            rd.cell(i, j).font = DATA_FONT
    rd.freeze_panes = "A2"

    # ── Sammendrag — totaler per sort/dag/måned/år, med grafer ──────────────
    # Alle tall er formler (COUNTIFS/SUMIFS) mot Rådata-arket, så de
    # oppdateres om noen redigerer/filtrerer bort rader der.
    sm = wb.create_sheet("Summary", 0)
    sm.sheet_view.showGridLines = False

    for j, br in enumerate([16, 10, 10, 15, 13, 12, 12, 12, 12, 12], start=1):
        sm.column_dimensions[get_column_letter(j)].width = br

    # Kompakt grønt banner (rad 1–3) med hvit knockout-logo integrert til venstre
    # og rapporttittelen til høyre — samme visuelle sprawk som fane-titlene, og
    # tar bare ~55 px høyde i stedet for en halv skjerm.
    BANNER_H = [20, 20, 15]          # radhøyder i banneret
    for i, h in enumerate(BANNER_H, start=1):
        sm.row_dimensions[i].height = h
        for kol_i in range(1, 11):
            sm.cell(i, kol_i).fill = TITTEL_FILL

    logo_sti = Path(__file__).parent / "skaak_logo_hvit.png"
    if logo_sti.exists():
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        with PILImage.open(logo_sti) as _im:
            fw, fh = _im.size
        vis_h = 40                    # px — passer inni banneret
        logo = XLImage(str(logo_sti))
        logo.height, logo.width = vis_h, int(vis_h * fw / fh)
        logo.anchor = "A1"
        sm.add_image(logo, "A1")

    sm.merge_cells("D1:J2")
    c = sm.cell(1, 4, "Production control — Half-year report")
    c.font = Font(name=FONT_NAVN, size=13, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    sm.merge_cells("D3:J3")
    c = sm.cell(3, 4, f"Generated {datetime.datetime.now():%d.%m.%Y %H:%M}  ·  "
                      f"{len(ok_rader)} packages  ·  formulas over Raw data sheet")
    c.font = Font(name=FONT_NAVN, size=8, italic=True, color="D9E5E1")
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    RD = "'Raw data'"
    MND_NAVN = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    kategorier = [k for k in FANE_REKKEFØLGE
                  if any(r["kategori"] == k for r in ok_rader)]
    år_liste = sorted({r["dato"].year for r in ok_rader})
    mnd_liste = sorted({(r["dato"].year, r["dato"].month) for r in ok_rader})
    dag_liste = sorted({r["dato"] for r in ok_rader})[-21:]

    def sm_overskrift(rad, titler):
        for j, t in enumerate(titler, start=1):
            c = sm.cell(rad, j, t)
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sm.row_dimensions[rad].height = 26

    def sm_seksjon(rad, tekst):
        c = sm.cell(rad, 1, tekst); c.font = SUBTOTAL_FONT

    def sm_celle(rad, kol, verdi, fmt=None, fet=False):
        c = sm.cell(rad, kol, verdi)
        c.font = Font(name=FONT_NAVN, size=10, bold=fet,
                      color="1B4D3E" if fet else "000000")
        c.border = RAMME
        if fmt: c.number_format = fmt
        if fet: c.fill = SUBTOTAL_FILL
        return c

    def cnt(krit):  return f"COUNTIFS({krit})"
    def sums(kol, krit): return f"SUMIFS({RD}!${kol}:${kol},{krit})"
    def k_kat(kat):      return f"{RD}!$B:$B,\"{kat}\""
    def k_år(y):         return (f"{RD}!$A:$A,\">=\"&DATE({y},1,1),"
                                 f"{RD}!$A:$A,\"<\"&DATE({y+1},1,1)")
    def k_mnd(y, m):
        y2, m2 = (y + 1, 1) if m == 12 else (y, m + 1)
        return (f"{RD}!$A:$A,\">=\"&DATE({y},{m},1),"
                f"{RD}!$A:$A,\"<\"&DATE({y2},{m2},1)")
    def k_dag(d):        return f"{RD}!$A:$A,DATE({d.year},{d.month},{d.day})"

    STD = ["Packages", "Boards", "Running metres (lm)", "Volume (m³)"]
    STD_FMT = ["#,##0", "#,##0", "#,##0.0", "0.000"]

    def std_formler(krit):
        return [f"={cnt(krit)}", f"={sums('E', krit)}",
                f"={sums('F', krit)}", f"={sums('G', krit)}"]

    # 1) Totalt per sortkategori ------------------------------------------------
    rad = 5
    sm_seksjon(rad, "Total per sort category"); rad += 1
    kat_hode = rad
    sm_overskrift(rad, ["Category"] + STD); rad += 1
    kat_forste = rad
    for kat in kategorier:
        sm_celle(rad, 1, kat)
        for j, (f, fmt) in enumerate(zip(std_formler(k_kat(kat)), STD_FMT), start=2):
            sm_celle(rad, j, f, fmt)
        rad += 1
    kat_siste = rad - 1
    sm_celle(rad, 1, "Total", fet=True)
    for j, fmt in enumerate(STD_FMT, start=2):
        kb = get_column_letter(j)
        sm_celle(rad, j, f"=SUM({kb}{kat_forste}:{kb}{kat_siste})", fmt, fet=True)
    rad += 2

    # 2) Per år -----------------------------------------------------------------
    sm_seksjon(rad, "Per year"); rad += 1
    sm_overskrift(rad, ["Year"] + STD); rad += 1
    for y in år_liste:
        sm_celle(rad, 1, str(y))
        for j, (f, fmt) in enumerate(zip(std_formler(k_år(y)), STD_FMT), start=2):
            sm_celle(rad, j, f, fmt)
        rad += 1
    rad += 1

    # 3) Per måned (+ kubikk per sortkategori, for stablet graf) ---------------
    sm_seksjon(rad, "Per month"); rad += 1
    mnd_hode = rad
    sm_overskrift(rad, ["Month"] + STD + [f"{k} (m³)" for k in kategorier]); rad += 1
    mnd_forste = rad
    for (y, m) in mnd_liste:
        sm_celle(rad, 1, f"{MND_NAVN[m]} {y}")
        for j, (f, fmt) in enumerate(zip(std_formler(k_mnd(y, m)), STD_FMT), start=2):
            sm_celle(rad, j, f, fmt)
        for j, kat in enumerate(kategorier, start=6):
            sm_celle(rad, j, f"={sums('G', k_mnd(y, m) + ',' + k_kat(kat))}", "0.000")
        rad += 1
    mnd_siste = rad - 1
    rad += 1

    # 4) Per dag (siste 21 produksjonsdager) -----------------------------------
    sm_seksjon(rad, f"Per day (last {len(dag_liste)} production days)"); rad += 1
    dag_hode = rad
    sm_overskrift(rad, ["Date"] + STD); rad += 1
    dag_forste = rad
    for d in dag_liste:
        sm_celle(rad, 1, d, "DD.MM.YYYY")
        for j, (f, fmt) in enumerate(zip(std_formler(k_dag(d)), STD_FMT), start=2):
            sm_celle(rad, j, f, fmt)
        rad += 1
    dag_siste = rad - 1

    # Grafer -------------------------------------------------------------------
    def stil(ch, tittel, h=8.2, b=15.5):
        ch.title = tittel; ch.height = h; ch.width = b
        ch.style = 10
        return ch

    kake = stil(PieChart(), "Volume by sort category", h=7.6, b=11.5)
    kake.add_data(Reference(sm, min_col=5, min_row=kat_forste, max_row=kat_siste))
    kake.set_categories(Reference(sm, min_col=1, min_row=kat_forste, max_row=kat_siste))
    sm.add_chart(kake, "L4")

    stab = stil(BarChart(), "Volume per month, by sort")
    stab.type = "col"; stab.grouping = "stacked"; stab.overlap = 100
    stab.add_data(Reference(sm, min_col=6, max_col=5 + len(kategorier),
                            min_row=mnd_hode, max_row=mnd_siste), titles_from_data=True)
    stab.set_categories(Reference(sm, min_col=1, min_row=mnd_forste, max_row=mnd_siste))
    stab.y_axis.title = "m³"
    sm.add_chart(stab, "L20")

    meter = stil(BarChart(), "Running metres per month")
    meter.type = "col"
    meter.add_data(Reference(sm, min_col=4, min_row=mnd_hode, max_row=mnd_siste),
                   titles_from_data=True)
    meter.set_categories(Reference(sm, min_col=1, min_row=mnd_forste, max_row=mnd_siste))
    meter.y_axis.title = "lm"; meter.legend = None
    sm.add_chart(meter, "L37")

    linje_ch = stil(LineChart(), f"Volume per day (last {len(dag_liste)} days)")
    linje_ch.add_data(Reference(sm, min_col=5, min_row=dag_hode, max_row=dag_siste),
                      titles_from_data=True)
    linje_ch.set_categories(Reference(sm, min_col=1, min_row=dag_forste, max_row=dag_siste))
    linje_ch.y_axis.title = "m³"; linje_ch.legend = None
    sm.add_chart(linje_ch, "L53")

    wb.active = 0

    tmp = xlsx_sti.with_suffix(".tmp.xlsx")
    wb.save(tmp); os.replace(tmp, xlsx_sti)
    antall = {fane: len(rader) for fane, rader in fane_data.items() if rader}
    log(f"Exported to {xlsx_sti}  (" + ", ".join(f"{f}: {n}" for f, n in antall.items()) + ")")



def main():
    p = argparse.ArgumentParser(description="rs232excel - Package label capture — RS-232 → CSV/Excel.")
    p.add_argument("--port", default="/dev/ttyUSB0")
    p.add_argument("--baud", type=int, default=9600)
    p.add_argument("--databits", type=int, default=8)
    p.add_argument("--parity", choices=["N", "E", "O"], default="N")
    p.add_argument("--stoppbits", type=int, default=1)
    p.add_argument("--timeout", type=float, default=1.0)
    p.add_argument("--flush", type=float, default=3.0)
    p.add_argument("--reset-terskel", type=int, default=100,
                   help="hvor stort hopp bakover som regnes som nullstilling")
    p.add_argument("--csv", default="packages.csv")
    p.add_argument("--xlsx", default="pakkelapper.xlsx")
    p.add_argument("--utskrift", default="capture.txt")
    p.add_argument("--mangler", default="missing.csv")
    p.add_argument("--summary-fil", default="oppsummering.csv")
    p.add_argument("--dimensjon-fil", default="oppsummering_dimensjon.csv")
    p.add_argument("--sesong-fil", default="season.txt")
    p.add_argument("--raw-capture", action="store_true")
    p.add_argument("--simulate")
    p.add_argument("--list-porter", action="store_true")
    p.add_argument("--export-xlsx", action="store_true")
    p.add_argument("--summary", action="store_true")
    p.add_argument("--summary-dimensjon", action="store_true",
                   help="antall/plank/kubikk gruppert per dimensjon (uavhengig av dato)")
    p.add_argument("--register", help="registrer et pakkenr manuelt (når lappen aldri kom)")
    p.add_argument("--set-season", help='sett gjeldende sesong: "raw" eller "kiln-dried"')
    args = p.parse_args()

    utskrift, csv_sti = Path(args.utskrift), Path(args.csv)
    mangler_sti, xlsx_sti, sesong_fil = Path(args.mangler), Path(args.xlsx), Path(args.sesong_fil)

    if args.list_porter:    list_porter(); return
    if args.set_season:    set_season(args.set_season, sesong_fil); return
    if args.export_xlsx: eksporter_xlsx(csv_sti, xlsx_sti); return
    if args.summary:   oppsummering(csv_sti, Path(args.summary_fil)); return
    if args.summary_dimensjon: oppsummering_dimensjon(csv_sti, Path(args.dimensjon_fil)); return

    sesong = read_season(sesong_fil)
    if args.register:
        registrer_manuelt(args.register, csv_sti, mangler_sti, sesong, args.reset_terskel); return

    reg = Register(csv_sti, mangler_sti, args.reset_terskel)
    if args.simulate: kjor_simulering(args, utskrift, csv_sti, reg, sesong)
    else:            les_serie(args, utskrift, csv_sti, reg, sesong)


if __name__ == "__main__":
    main()
